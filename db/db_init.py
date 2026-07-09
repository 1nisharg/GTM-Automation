"""
db/db_init.py
-------------
Auto-migration and auto-seed runner.
Called once at server startup from main.py lifespan.

What it does:
1. Runs init.sql        — creates partners table + indexes
2. Runs db_schema.sql   — creates voice agent tables
3. Checks if Excel seed is needed — seeds partners if:
   a. partners table is empty, OR
   b. Excel file is newer than last_seeded_at in a metadata table

The client never needs to touch Supabase. Just update the Excel and restart.
"""

import hashlib
import logging
import os
from pathlib import Path

from db.connection import get_pool

logger = logging.getLogger(__name__)

# Paths — relative to project root
_ROOT        = Path(__file__).resolve().parent.parent
_INIT_SQL    = _ROOT / "db" / "init.sql"
_SCHEMA_SQL  = _ROOT / "voice_agent" / "db_schema.sql"
_EXCEL_FILE  = _ROOT / "GTM UAE_ Track 1 & 2 Db.xlsx"

TARGET_STATUSES = {"Partner Outreach", "Yet to Start"}


# ── Helpers ────────────────────────────────────────────────────────────────────

async def _run_sql_file(conn, path: Path) -> None:
    """Execute a .sql file against the DB connection."""
    if not path.exists():
        logger.warning("db_init: SQL file not found: %s", path)
        return
    sql = path.read_text(encoding="utf-8")
    await conn.execute(sql)
    logger.info("db_init: ran %s", path.name)


async def _ensure_meta_table(conn) -> None:
    """Create a lightweight metadata table to track seed state."""
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS _seed_meta (
            key   TEXT PRIMARY KEY,
            value TEXT
        )
    """)


async def _ensure_api_usage_table(conn) -> None:
    """
    Create the api_usage table — tracks Hunter (and any future non-Apollo
    API) usage for daily-limit enforcement and audit.

    This is embedded directly here (rather than a separate .sql file) so
    it can never silently go missing again — a prior version of this
    project referenced an external api_usage_schema.sql file that was
    lost during a refactor, which caused Hunter's daily-limit check and
    usage logging to fail silently (caught by try/except, defaulting to
    "allow the call") with zero visible error until a manual DB reset
    surfaced "relation api_usage does not exist".

    Apollo has its own separate apollo_usage table (created by init.sql)
    with its own schema — this table is for Hunter/Tavily/other sources.
    """
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS api_usage (
            id           SERIAL PRIMARY KEY,
            run_id       TEXT DEFAULT '',
            partner_name TEXT DEFAULT '',
            api_name     TEXT NOT NULL,
            operation    TEXT NOT NULL,
            success      BOOLEAN DEFAULT true,
            result       TEXT DEFAULT '',
            request_cost NUMERIC DEFAULT 1,
            called_at    TIMESTAMP DEFAULT now()
        )
    """)
    await conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_api_usage_api_name  ON api_usage (api_name)
    """)
    await conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_api_usage_run_id    ON api_usage (run_id)
    """)
    await conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_api_usage_called_at ON api_usage (called_at DESC)
    """)


async def _get_meta(conn, key: str) -> str | None:
    row = await conn.fetchrow("SELECT value FROM _seed_meta WHERE key = $1", key)
    return row["value"] if row else None


async def _set_meta(conn, key: str, value: str) -> None:
    await conn.execute("""
        INSERT INTO _seed_meta (key, value) VALUES ($1, $2)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
    """, key, value)


def _excel_hash() -> str | None:
    """Return MD5 hash of the Excel file, or None if file doesn't exist."""
    if not _EXCEL_FILE.exists():
        return None
    h = hashlib.md5()
    with open(_EXCEL_FILE, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _parse_excel() -> list[tuple]:
    """Parse Excel and return deduplicated partner rows matching target statuses."""
    try:
        import openpyxl
    except ImportError:
        logger.error("db_init: openpyxl not installed — cannot seed from Excel")
        return []

    if not _EXCEL_FILE.exists():
        logger.warning("db_init: Excel file not found at %s", _EXCEL_FILE)
        return []

    wb = openpyxl.load_workbook(_EXCEL_FILE, read_only=True)
    records = {}  # key=(name, sheet_source) for dedup

    def clean(val):
        if val is None:
            return None
        s = str(val).strip()
        if s.startswith("=+"):
            s = s[1:]
        return s if s and s != "None" else None

    def safe_int(val):
        try:
            return int(float(val))
        except:
            return None

    # Track 1
    ws1 = wb["Track 1 Db"]
    for row in list(ws1.iter_rows(values_only=True))[1:]:
        status = clean(row[7])
        if status not in TARGET_STATUSES:
            continue
        name = clean(row[1])
        if not name:
            continue
        records[(name, "track1")] = (
            name, clean(row[2]), clean(row[3]), clean(row[4]),
            clean(row[5]), safe_int(row[6]), status, clean(row[8]),
            clean(row[9]), clean(row[10]), clean(row[11]), clean(row[12]), "track1"
        )

    # Track 2
    ws2 = wb["Track 2 Db"]
    for row in list(ws2.iter_rows(values_only=True))[1:]:
        status = clean(row[7])
        if status not in TARGET_STATUSES:
            continue
        name = clean(row[1])
        if not name:
            continue
        records[(name, "track2")] = (
            name, clean(row[2]), clean(row[3]), clean(row[4]),
            clean(row[5]), safe_int(row[6]), status, clean(row[8]),
            clean(row[9]), clean(row[10]), clean(row[11]), clean(row[12]), "track2"
        )

    logger.info("db_init: parsed %d unique partners from Excel", len(records))
    return list(records.values())


async def _seed_partners(conn, records: list[tuple]) -> None:
    """Upsert all partner records into the partners table."""
    if not records:
        return

    # Batch upsert in chunks of 200 to avoid query size limits
    chunk_size = 200
    total = 0
    for i in range(0, len(records), chunk_size):
        chunk = records[i:i + chunk_size]
        # Insert with NULL contact fields — we deliberately do NOT seed
        # phone_number, email_id, linkedin_profile from Excel because:
        # 1. The Excel data is often generic (info@, outdated numbers)
        # 2. The enrichment pipeline will find real personal contacts
        # 3. After a reset we want a fully clean enrichment slate
        # Strip contact columns from each record before insert:
        clean_chunk = [
            (r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], None, None, None, r[12])
            for r in chunk
        ]
        await conn.executemany("""
            INSERT INTO partners
                (partner_name, digitisation, category, subcategories, website,
                 product_count, status, integrated, region,
                 phone_number, email_id, linkedin_profile, sheet_source)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)
            ON CONFLICT (partner_name, sheet_source) DO UPDATE SET
                digitisation     = EXCLUDED.digitisation,
                category         = EXCLUDED.category,
                subcategories    = EXCLUDED.subcategories,
                website          = EXCLUDED.website,
                product_count    = EXCLUDED.product_count,
                status           = EXCLUDED.status,
                integrated       = EXCLUDED.integrated,
                region           = EXCLUDED.region
        """, clean_chunk)
        total += len(chunk)

    logger.info("db_init: upserted %d partners into Supabase", total)


# ── Main entry point ───────────────────────────────────────────────────────────

async def run_startup_migrations() -> None:
    """
    Run all migrations and conditional seeding at server startup.
    Safe to call every time — all operations are idempotent.
    """
    logger.info("db_init: running startup migrations…")
    pool = await get_pool()

    async with pool.acquire() as conn:

        # ── Step 1: Run init.sql (partners table + indexes) ────────────────
        try:
            await _run_sql_file(conn, _INIT_SQL)
        except Exception as exc:
            logger.warning("db_init: init.sql error (table may already exist): %s", exc)

        # ── Step 2: Run voice agent schema ─────────────────────────────────
        try:
            await _run_sql_file(conn, _SCHEMA_SQL)
        except Exception as exc:
            logger.warning("db_init: db_schema.sql error (tables may already exist): %s", exc)

        # ── Step 2b: Re-run init.sql to pick up outreach_sequence table ────
        # Uses CREATE TABLE IF NOT EXISTS throughout — safe to re-run every boot.
        try:
            await _run_sql_file(conn, _INIT_SQL)
        except Exception as exc:
            logger.warning("db_init: init.sql re-run error: %s", exc)

        # ── Step 3: Ensure unique constraint for upsert ────────────────────
        try:
            await conn.execute("""
                ALTER TABLE partners
                ADD CONSTRAINT uq_partner_name_source
                UNIQUE (partner_name, sheet_source)
            """)
            logger.info("db_init: unique constraint added")
        except Exception:
            pass  # Already exists — fine

        # ── Step 4: Meta table for seed tracking ───────────────────────────
        try:
            await _ensure_meta_table(conn)
        except Exception as exc:
            logger.warning("db_init: meta table error: %s", exc)

        # ── Step 4b: api_usage table for Hunter/Tavily daily-limit tracking ─
        # This was previously an external .sql file that silently went
        # missing during a refactor — now embedded directly so it can
        # never be dropped again without a visible code change.
        try:
            await _ensure_api_usage_table(conn)
            logger.info("db_init: api_usage table ready.")
        except Exception as exc:
            logger.warning("db_init: api_usage table creation error: %s", exc)

        # ── Step 5: Check if Excel seed needed ─────────────────────────────
        try:
            # Log exact path so startup logs show whether file is found
            logger.info("db_init: Excel path = %s (exists=%s)", _EXCEL_FILE, _EXCEL_FILE.exists())

            current_hash  = _excel_hash()
            stored_hash   = await _get_meta(conn, "excel_hash")
            partner_count = await conn.fetchval("SELECT COUNT(*) FROM partners")

            logger.info(
                "db_init: partner_count=%d excel_found=%s hash_match=%s",
                partner_count, _EXCEL_FILE.exists(),
                current_hash == stored_hash if current_hash else "no_hash",
            )

            needs_seed = (
                current_hash is not None and (
                    partner_count == 0 or
                    current_hash != stored_hash
                )
            )

            if needs_seed:
                reason = "table empty" if partner_count == 0 else "Excel file changed"
                logger.info("db_init: seeding partners (%s)…", reason)
                records = _parse_excel()
                if not records:
                    logger.error(
                        "db_init: Excel parsed 0 records — check file exists at %s "
                        "and sheet names are 'Track 1 Db' / 'Track 2 Db'", _EXCEL_FILE
                    )
                else:
                    # Seed using the existing conn — mark meta immediately after
                    await _seed_partners(conn, records)
                    if current_hash:
                        await _set_meta(conn, "excel_hash", current_hash)
                    logger.info("db_init: seed complete — %d partners in DB", len(records))
            else:
                logger.info(
                    "db_init: partners table has %d rows, Excel unchanged — skipping seed.",
                    partner_count,
                )
        except Exception as exc:
            logger.error("db_init: seed check failed: %s", exc)

    # ── Step 6: Populate subcategory_tags from subcategories field ───────
    try:
        async with pool.acquire() as conn:
            result = await conn.execute("""
                UPDATE partners
                SET subcategory_tags = (
                    SELECT ARRAY_AGG(TRIM(tag))
                    FROM UNNEST(string_to_array(subcategories, ',')) AS tag
                    WHERE TRIM(tag) != ''
                )
                WHERE subcategories IS NOT NULL
                  AND (subcategory_tags IS NULL OR array_length(subcategory_tags, 1) IS NULL)
            """)
            logger.info("db_init: subcategory_tags populated — %s", result)
    except Exception as exc:
        logger.warning("db_init: subcategory_tags population failed: %s", exc)

    logger.info("db_init: startup migrations complete.")