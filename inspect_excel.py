import openpyxl

wb = openpyxl.load_workbook("GTM UAE_ Track 1 & 2 Db.xlsx", read_only=True)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n=== Sheet: {sheetname} ===")
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        print("Headers:", rows[0])
        print("Row count (approx):", len(rows))
        print("Sample rows (first 5 data rows):")
        for r in rows[1:6]:
            print(" ", r)
        # Print unique subcategory values
        subcategory_col = None
        status_col = None
        if rows:
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            for i, h in enumerate(headers):
                if "subcategory" in h or "sub_category" in h or "sub category" in h:
                    subcategory_col = i
                if "status" in h or "enrichment" in h:
                    status_col = i
        if subcategory_col is not None:
            cats = set()
            for r in rows[1:]:
                if r[subcategory_col]:
                    cats.add(str(r[subcategory_col]).strip())
            print(f"Unique subcategory values ({len(cats)}):", sorted(cats)[:20])
        if status_col is not None:
            statuses = set()
            for r in rows[1:]:
                if r[status_col]:
                    statuses.add(str(r[status_col]).strip())
            print(f"Status column (col {status_col}) unique values:", statuses)
