"""
scripts/export_partners_to_xlsx.py
Export the partners table from Supabase to a formatted Excel file.

Usage (from project root):
    python scripts/export_partners_to_xlsx.py

Output:
    partners_export.xlsx
"""

import asyncio
import os
import ssl
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import asyncpg
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

for candidate in [Path(".env"), Path("../.env"), Path(__file__).parent.parent / ".env"]:
    if candidate.exists():
        load_dotenv(candidate)
        break

DATABASE_URL = os.getenv("DATABASE_URL", "")
if not DATABASE_URL:
    print("ERROR: DATABASE_URL not found in .env")
    sys.exit(1)

# Exact columns from your partners table (in display order)
COLUMNS = [
    # (db_column,          display header,              width)
    ("id",                 "ID",                        7),
    ("partner_name",       "Partner Name",              36),
    ("digitisation",       "Digitisation",              18),
    ("category",           "Category",                  22),
    ("subcategories",      "Subcategories",             40),
    ("subcategory_tags",   "Subcategory Tags",          40),
    ("region",             "Region",                    16),
    ("status",             "Status",                    18),
    ("integrated",         "Integrated",                12),
    ("product_count",      "Product Count",             14),
    ("website",            "Website",                   30),
    ("phone_number",       "Phone Number",              20),
    ("email_id",           "Email",                     30),
    ("linkedin_profile",   "LinkedIn Profile",          36),
    ("sheet_source",       "Sheet Source",              14),
]

DB_COLS  = [c[0] for c in COLUMNS]
HEADERS  = [c[1] for c in COLUMNS]
WIDTHS   = [c[2] for c in COLUMNS]


async def fetch_partners():
    print("Connecting to Supabase...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode    = ssl.CERT_NONE

    conn = await asyncpg.connect(DATABASE_URL, ssl=ctx, statement_cache_size=0)

    # Only select columns that actually exist in the table
    existing = await conn.fetch("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name = 'partners' AND table_schema = 'public'
    """)
    existing_set = {r["column_name"] for r in existing}

    safe_cols = [c for c in DB_COLS if c in existing_set]
    missing   = [c for c in DB_COLS if c not in existing_set]
    if missing:
        print(f"Note: columns not in DB (skipped): {missing}")

    col_sql = ", ".join(f'"{c}"' for c in safe_cols)
    rows    = await conn.fetch(f'SELECT {col_sql} FROM partners ORDER BY id')
    await conn.close()

    print(f"Fetched {len(rows):,} rows, {len(safe_cols)} columns.")
    return safe_cols, [dict(r) for r in rows]


def build_workbook(safe_cols: list, rows: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partners"

    # Rebuild column config for only the cols that exist
    col_cfg = [(db, hdr, w) for db, hdr, w in COLUMNS if db in safe_cols]

    # ── Styles ─────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    even_fill  = PatternFill("solid", fgColor="DCE6F1")
    odd_fill   = PatternFill("solid", fgColor="FFFFFF")
    body_font  = Font(name="Arial", size=10)
    body_align = Alignment(vertical="center")

    digit_color = {
        "Fully digitised": "C6EFCE",
        "Semi-digitised":  "FFEB9C",
        "Un-digitised":    "FFC7CE",
    }
    status_color = {
        "Live":              "C6EFCE",
        "Partner Outreach":  "DDEBF7",
        "Yet to Start":      "FFC7CE",
        "In Progress":       "FFEB9C",
    }

    # ── Header ─────────────────────────────────────────────────────────────
    for ci, (_, hdr, _w) in enumerate(col_cfg, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[1].height = 30

    # Column index lookups for conditional colouring
    col_pos = {db: ci for ci, (db, _, _) in enumerate(col_cfg, 1)}

    # ── Data rows ──────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, 2):
        base = even_fill if ri % 2 == 0 else odd_fill
        for ci, (db, _, _) in enumerate(col_cfg, 1):
            val = row.get(db)
            if val is None:
                val = ""
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = body_font
            cell.fill      = base
            cell.alignment = body_align

        # Conditional cell colours
        if "digitisation" in col_pos:
            dv = row.get("digitisation") or ""
            if dv in digit_color:
                ws.cell(row=ri, column=col_pos["digitisation"]).fill = PatternFill(
                    "solid", fgColor=digit_color[dv])

        if "status" in col_pos:
            sv = row.get("status") or ""
            if sv in status_color:
                ws.cell(row=ri, column=col_pos["status"]).fill = PatternFill(
                    "solid", fgColor=status_color[sv])

    # ── Column widths, freeze, filter ──────────────────────────────────────
    for ci, (_, _, w) in enumerate(col_cfg, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    last_col = get_column_letter(len(col_cfg))
    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Aarna UAE Partners — Export Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13)

    ws2["A3"] = "Total partners"
    ws2["B3"] = len(rows)
    ws2["B3"].font = Font(name="Arial", bold=True, size=12)

    ws2["A4"] = "Exported at"
    ws2["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    def write_breakdown(start_row, title, field):
        ws2.cell(row=start_row, column=1, value=title).font = Font(name="Arial", bold=True)
        counts = Counter(r.get(field) or "Unknown" for r in rows)
        for i, (k, v) in enumerate(sorted(counts.items()), start_row + 1):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=v)
        return start_row + len(counts) + 2

    next_row = 6
    if "digitisation" in safe_cols:
        next_row = write_breakdown(next_row, "Digitisation Breakdown", "digitisation")
    if "status" in safe_cols:
        next_row = write_breakdown(next_row, "Status Breakdown", "status")
    if "category" in safe_cols:
        next_row = write_breakdown(next_row, "Category Breakdown", "category")
    if "region" in safe_cols:
        next_row = write_breakdown(next_row, "Region Breakdown", "region")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16

    return wb


async def main():
    safe_cols, rows = await fetch_partners()

    print("Building Excel file...")
    wb = build_workbook(safe_cols, rows)

    out = Path("partners_export.xlsx")
    wb.save(out)
    kb = out.stat().st_size // 1024
    print(f"\nSaved : {out.resolve()}")
    print(f"Rows  : {len(rows):,}")
    print(f"Size  : {kb} KB")
    print("\nUpload to Google Drive → right-click → Open with Google Sheets")


if __name__ == "__main__":
    asyncio.run(main())